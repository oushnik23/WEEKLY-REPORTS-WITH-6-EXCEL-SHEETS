import pandas as pd
import datetime
import numpy as np
import mysql.connector as msql
from mysql.connector import Error
from google.cloud import bigquery
import warnings
warnings.filterwarnings("ignore")
import math

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

import os 
os.chdir(r"D:\Oushnik Sarkar\Python\Weekly\GGL")
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = 'D:/Oushnik Sarkar/data-warehousing-prod.json'

#Imports google cloud client library and initiates BQ service
from google.cloud import bigquery
#from google.cloud import datastore
bigquery_client = bigquery.Client()

QUERY = """
With t1 as(
select FinYear,
CASE 
    WHEN Area = "AS" AND Centre IN ("KOL", "GUW") THEN "AS"
    WHEN Area IN ("DO", "TR") AND Centre IN ("KOL", "SIL") THEN "DO/TR"
    WHEN Area IN ("CA", "TP") AND Centre IN ("KOL", "GUW") THEN "CA/TP" ELSE Area END AS AreaAlies,


if(SaleNo >=1 and SaleNo <=13, 52+SaleNo,SaleNo) as SalesAlies,
BrokerCode,Category,GardenMDM,MDMGradeGroup,
coalesce(Centre,"") as Centre,SellerGroup,
coalesce(EstBlf,"") as EstBlf, 
ROUND( SAFE_DIVIDE(SUM(Value) , SUM(TotalWeight)) ,2) AS AvgPrice,
coalesce(sum(IF(LotStatus = 'Sold',TotalWeight,InvoiceWeight)),0) as OfferQty,
sum(TotalWeight) as Sold_Qty,
coalesce(sum(Value),0) as Total_Value

FROM `data-warehousing-prod.EasyReports.SaleTransactionView`

where Area IN ("AS") and EstBlf IN ("EST") and Centre IN ("KOL","GUW") and Season IN (2025) and Category IN ("ORTHODOX") and
if(SaleNo >=1 and SaleNo <=13, 52+SaleNo,SaleNo) =
(Select max(if(SaleNo>=1 AND SaleNo<=13,52+SaleNo,SaleNo)) from data-warehousing-prod.EasyReports.SaleTransactionView 
where Season = 2025 and FinYear = "2025-26")

group by FinYear,SalesAlies,BrokerCode,Category,GardenMDM,MDMGradeGroup,AreaAlies,Centre,SellerGroup,EstBlf
Having FinYear IS NOT NULL),
t2 as(
Select FinYear,Season,GardenMDM,

coalesce( ROUND(sum(Value),0),0) as Total_Value,
 coalesce(ROUND(sum(TotalWeight),0),0) as Sold_Qty,
 case when sum(TotalWeight)>=5000 then dense_rank() OVER (Partition by concat(FinYear,Season) order by if(sum(TotalWeight)>=5000,SUM(Value) / SUM(TotalWeight),0) DESC) else 0 end AS BOP
 
 from `data-warehousing-prod.EasyReports.SaleTransactionView`
 
 Where Area IN ("AS") and EstBlf IN ("EST") and Centre IN ("KOL","GUW") and Season IN (2025) and Category IN ("ORTHODOX") and
if(SaleNo >=1 and SaleNo <=13, 52+SaleNo,SaleNo) =
(Select max(if(SaleNo>=1 AND SaleNo<=13,52+SaleNo,SaleNo)) from data-warehousing-prod.EasyReports.SaleTransactionView where Season = 2025 and FinYear = "2025-26")

Group By FinYear,Season,GardenMDM Having FinYear IS NOT NULL)

Select t1.FinYear,t1.AreaAlies, t1.SalesAlies, t1.BrokerCode, t1.Category, t1.MDMGradeGroup, t1.GardenMDM, t1.Centre, t1.SellerGroup, t1.EstBlf, t1.Sold_Qty,t1.OfferQty,t1.Total_Value,t1.AvgPrice,t2.BOP from t1 Left Join t2 ON

t1.FinYear=t2.FinYear And t1.GardenMDM=t2.GardenMDM
"""
Query_Results = bigquery_client.query(QUERY)
df = Query_Results.to_dataframe()

df1=df[(df['BOP']>=1) & (df['BOP']<=10)]

df2=df[(df['BOP']>=11) & (df['BOP']<=20)]

df1.columns

Sale = df['SalesAlies'].max() - 52 if df['SalesAlies'].max() > 52 else df['SalesAlies'].max()

#----------------------------------BOP 1-10----------------------------------#
bop1=df1[['GardenMDM','BOP']].drop_duplicates().reset_index(drop=True).sort_values(by='BOP', ascending=True)
garden_order1 = bop1['GardenMDM'].tolist()

summary1=df1.groupby(['MDMGradeGroup','GardenMDM'],as_index=False).agg({'Sold_Qty':'sum','Total_Value':'sum'})

summary1['Avg']=summary1['Total_Value'] / summary1['Sold_Qty']
summary1.drop(['Total_Value'],inplace=True,axis=1)

dfa=summary1.set_index(['MDMGradeGroup','GardenMDM']).unstack(['GardenMDM'])
dfa=dfa.swaplevel(axis=1).sort_index(axis=1)

#--------------------------------------SUBTOTAL2---------------------------------------------------#

sold_qty = dfa.xs('Sold_Qty', axis=1, level=1)
avg = dfa.xs('Avg', axis=1, level=1)

subtotal_qty = sold_qty.sum()
weighted_avg = (sold_qty * avg).sum() / subtotal_qty.replace(0, np.nan)

# Step 2: Build MultiIndex columns like df11
gardens = subtotal_qty.index
columns = pd.MultiIndex.from_product(
    [gardens, ['Sold_Qty', 'Avg']],
    names=['GardenMDM', 'Metric']
)

# Step 3: Build values in same order as columns
values = []
for g in gardens:
    values.extend([subtotal_qty[g], weighted_avg[g]])

# Step 4: Create the DataFrame with one row
Subtotal1 = pd.DataFrame([values], columns=columns)

# Step 5: Set the MultiIndex row for Grand Total
Subtotal1.index = pd.Index(['Grand Total'], name=dfa.index.name)

# Step 6: Append to df11
dfa2 = pd.concat([dfa, Subtotal1], axis=0)

#--------------------------------------Desired Orders---------------------------------------------------#

desired_order=['Sold_Qty','Avg']
new_columns = []
for i in dfa2.columns.levels[0]: 
    for j in desired_order:  
        new_columns.append((i,j))
dfa2 = dfa2[new_columns]

grade_sequence = [
"1st LINE W.LEAF", "2nd LINE W.LEAF", "GFOP", "FOP", "OP", "OPA", "BPS", "FBOP", "GFBOP", "GBOP", "FANNINGS", "SECONDARIES", "ORTHODOX DUST"]

grade_order = {g: i for i, g in enumerate(grade_sequence)}

# Convert index to DataFrame
idx_df1 = dfa2.index.to_frame(index=False)

# Map sorting ranks
idx_df1['grade_rank'] = idx_df1['MDMGradeGroup'].map(grade_order).fillna(998)

# Force 'Total' to the end
idx_df1.loc[idx_df1['MDMGradeGroup'] == 'Grand Total', 'grade_rank'] = 999

# Sort and rebuild MultiIndex
sorted_index = idx_df1.sort_values(['grade_rank']) \
                     .drop(columns=['grade_rank']) \
                     .apply(tuple, axis=1)
                     
sorted_index = idx_df1.sort_values(['grade_rank'])['MDMGradeGroup'].tolist()

dfa2 = dfa2.loc[[g for g in sorted_index if g in dfa2.index]]
dfa2 = dfa2.reindex(columns=garden_order1, level=0)
dfa2.rename(columns={'Sold_Qty':'Sold Qty'},inplace=True)

dfa2=dfa2.reset_index()

dfa2.columns = pd.MultiIndex.from_tuples(
    [('', 'Grade')] + list(dfa2.columns[1:]))

#----------------------------------BOP 11-20----------------------------------#

bop2=df2[['GardenMDM','BOP']].drop_duplicates().reset_index(drop=True).sort_values(by='BOP', ascending=True)
garden_order2 = bop2['GardenMDM'].tolist()

summary2=df2.groupby(['MDMGradeGroup','GardenMDM'],as_index=False).agg({'Sold_Qty':'sum','Total_Value':'sum'})

summary2['Avg']=summary2['Total_Value'] / summary2['Sold_Qty']
summary2.drop(['Total_Value'],inplace=True,axis=1)

dfb=summary2.set_index(['MDMGradeGroup','GardenMDM']).unstack(['GardenMDM'])
dfb=dfb.swaplevel(axis=1).sort_index(axis=1)

#--------------------------------------SUBTOTAL2---------------------------------------------------#

sold_qty = dfb.xs('Sold_Qty', axis=1, level=1)
avg = dfb.xs('Avg', axis=1, level=1)

subtotal_qty = sold_qty.sum()
weighted_avg = (sold_qty * avg).sum() / subtotal_qty.replace(0, np.nan)

# Step 2: Build MultiIndex columns like df11
gardens = subtotal_qty.index
columns = pd.MultiIndex.from_product(
    [gardens, ['Sold_Qty', 'Avg']],
    names=['GardenMDM', 'Metric']
)

# Step 3: Build values in same order as columns
values = []
for g in gardens:
    values.extend([subtotal_qty[g], weighted_avg[g]])

# Step 4: Create the DataFrame with one row
Subtotal1 = pd.DataFrame([values], columns=columns)

# Step 5: Set the MultiIndex row for Grand Total
Subtotal1.index = pd.Index(['Grand Total'], name=dfb.index.name)

# Step 6: Append to df11
dfb2 = pd.concat([dfb, Subtotal1], axis=0)

#--------------------------------------Desired Orders---------------------------------------------------#

desired_order=['Sold_Qty','Avg']
new_columns = []
for i in dfb2.columns.levels[0]: 
    for j in desired_order:  
        new_columns.append((i,j))
dfb2 = dfb2[new_columns]

grade_sequence = [
"1st LINE W.LEAF", "2nd LINE W.LEAF", "GFOP", "FOP", "OP", "OPA", "BPS", "FBOP", "GFBOP", "GBOP", "FANNINGS", "SECONDARIES", "ORTHODOX DUST"]

grade_order = {g: i for i, g in enumerate(grade_sequence)}

# Convert index to DataFrame
idx_df2 = dfb2.index.to_frame(index=False)

# Map sorting ranks
idx_df2['grade_rank'] = idx_df2['MDMGradeGroup'].map(grade_order).fillna(998)

# Force 'Total' to the end
idx_df2.loc[idx_df2['MDMGradeGroup'] == 'Grand Total', 'grade_rank'] = 999

# Sort and rebuild MultiIndex
sorted_index = idx_df2.sort_values(['grade_rank']) \
                     .drop(columns=['grade_rank']) \
                     .apply(tuple, axis=1)
                     
sorted_index = idx_df2.sort_values(['grade_rank'])['MDMGradeGroup'].tolist()

dfb2 = dfb2.loc[[g for g in sorted_index if g in dfb2.index]]

dfb2 = dfb2.reindex(columns=garden_order2, level=0)
dfb2.rename(columns={'Sold_Qty':'Sold Qty'},inplace=True)

dfb2=dfb2.reset_index()

dfb2.columns = pd.MultiIndex.from_tuples(
    [('', 'Grade')] + list(dfb2.columns[1:]))

#--------------------------------------------DESIGN-------------------------------------------#

excel_path = "AS_EST_ORTH.xlsx"

with pd.ExcelWriter(excel_path, engine="xlsxwriter") as writer:
    dfa2.to_excel(writer, sheet_name="AS EST TOP 20 ORTH", startrow=3, index=True, header=True)
    workbook  = writer.book
    worksheet = writer.sheets["AS EST TOP 20 ORTH"]

    # ✅ Hide column A here
    worksheet.set_column('A:A', None, None, {'hidden': True})

# Step 2: Delete row(s) if needed before adding dfb2
wb = load_workbook(excel_path)
ws = wb["AS EST TOP 20 ORTH"]
ws.delete_rows(6)   # <-- Delete row 6 before appending dfb2

wb.save(excel_path)
wb.close()

# Step 3: Append dfb2 below dfa2
with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    start_row = 3 + len(dfa2) + dfa2.columns.nlevels + 2   # keep spacing
    dfb2.to_excel(writer, sheet_name="AS EST TOP 20 ORTH", startrow=start_row)

# Step 4: Format with openpyxl
wb = load_workbook(excel_path)
ws = wb["AS EST TOP 20 ORTH"]
#ws.delete_rows(28)

dfa2_end_row = 4 + len(dfa2) + dfa2.columns.nlevels   # last row of dfa2 (including headers)
delete_row = dfa2_end_row + 4
ws.delete_rows(delete_row)

thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),
                     bottom=Side(style='thin'))

dfa2_start_row = 4  # because you wrote it at startrow=3 and header takes 1 row
dfa2_end_row = dfa2_start_row + len(dfa2) + dfa2.columns.nlevels-1
dfa2_end_col = dfa2.shape[1] + (1 if dfa2.index.nlevels > 0 else 0)

# --- define boundaries for dfb2 ---
dfb2_start_row = start_row + 1  # Excel is 1-based
dfb2_end_row = dfb2_start_row + len(dfb2) + dfb2.columns.nlevels -1
dfb2_end_col = dfb2.shape[1] + (1 if dfb2.index.nlevels > 0 else 0)

# --- apply border only to dfa2 ---
for row in ws.iter_rows(min_row=dfa2_start_row, max_row=dfa2_end_row, max_col=dfa2_end_col):
    for cell in row:
        cell.border = thin_border

# --- apply border only to dfb2 ---
for row in ws.iter_rows(min_row=dfb2_start_row, max_row=dfb2_end_row, max_col=dfb2_end_col):
    for cell in row:
        cell.border = thin_border

ws["B1"] = "TOP 20 ASSAM EST ORTHODOX GARDEN"
ws["B2"] = f"FOR SALE - {Sale}"
ws["B3"] = "SEASON 2025/26"

bold_font = Font(bold=True,color="FF0000")
ws["B1"].font = bold_font
ws["B2"].font = bold_font
ws["B3"].font = bold_font

for col_letter in ["B", "C"]:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cell = row[ord(col_letter) - ord("A")]
        cell.alignment = Alignment(horizontal="left")
  
import re

for col in ws.iter_cols(min_col=3, max_col=ws.max_column, min_row=0, max_row=ws.max_row):
    for cell in col:
        if isinstance(cell.value, (int, float)) and cell.value == 0:
        # Option 1: Hide the 0 visually but keep value for formulas
            cell.number_format = ';;;'
    
    for cell in col:
        try:
            if isinstance(cell.value, str):
                
                cleaned = cell.value.replace(',', '').replace(' ', '')
                cleaned = re.sub(r'[^0-9.\-]', '', cell.value) 
                
                if re.match(r'^-?\d+(\.\d+)?$', cleaned): 
                    cell.value = float(cleaned)

                if cleaned.replace('.', '', 1).isdigit(): 
                    cell.value = float(cleaned) 
            
            if isinstance(cell.value, (int, float)):
                if cell.value == 0:
      # ✅ Hide zero values visually but keep them in formulas
                  cell.number_format = ';;;'
                else:
      # Normal formatting for non-zero numbers
                  cell.number_format = '#,##0;-#,##0'
                
                #Internally retain decimal value for future reference
                if isinstance(cell.value, float):
                    # Keep the original value (store as float internally)
                    pass  # No action needed, value remains as float

        except Exception as e:
            print(f"Skipping cell {cell.coordinate}: {e}")

# --- Auto adjust column width from D onwards (with max width limit) ---
for col in range(2, ws.max_column + 0):  # D = 4
    max_length = 0
    column_letter = get_column_letter(col)

    for cell in ws[column_letter]:
        try:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        except:
            pass

    # Add padding but cap at 40
    adjusted_width = min(max_length + 2, 10)
    ws.column_dimensions[column_letter].width = adjusted_width

targets = [('Grand Total')]

# Define the font style for Grand Total
highlight_font = Font(bold=True, color="974706")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    grade  = row[1].value  # Column D

    if (str(grade).strip().upper()) in \
       [g.upper() for g in targets]:
        for cell in row:
            cell.font = highlight_font
            
for col in range(3, ws.max_column + 1):
    for cell in ws.iter_cols(min_col=col, max_col=col, min_row=1, max_row=ws.max_row):
        for i in cell:
            i.alignment = Alignment(horizontal='center', vertical='center')

ws.sheet_view.zoomScale = 80
ws.column_dimensions['B'].width = 19
ws.freeze_panes = "C4"

footer_row = dfb2_start_row + len(dfb2) + dfb2.columns.nlevels  # 2 blank lines after dfb2
footer_cell = f"B{footer_row}"

ws[footer_cell] = "*B.O.P. CUT OFF 5000 KGS."

# Style for footer
footer_font = Font(bold=True, color="C00000")  # purple-like tone
ws[footer_cell].font = footer_font
ws[footer_cell].alignment = Alignment(horizontal="left", vertical="center")
                
wb.save(excel_path)
print(f"✅ Excel file {excel_path} created successfully!")
