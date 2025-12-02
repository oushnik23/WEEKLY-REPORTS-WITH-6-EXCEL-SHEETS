import os
import subprocess
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, PatternFill, Side
from openpyxl.utils import get_column_letter

# Define paths
working_directory = r"D:\Oushnik Sarkar\Python\Weekly\GGL"
os.chdir(working_directory)

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = 'D:/Oushnik Sarkar/data-warehousing-prod.json'

# Define script files and their expected output
scripts = [
    {"script": "AS CTC.py", "output": "AS CTC.xlsx"},
    {"script": "DO CTC.py", "output": "DO CTC.xlsx"},
    {"script": "AS ORTH.py", "output": "AS ORTH.xlsx"},
    {"script": "AS_EST_CTC.py", "output": "AS_EST_CTC.xlsx"},
    {"script": "DO_EST_CTC.py", "output": "DO_EST_CTC.xlsx"},
    {"script": "AS_EST_ORTH.py", "output": "AS_EST_ORTH.xlsx"},
]

output_file = "GGL FOR SALE- 48.xlsx"

def run_scripts_and_collect_outputs(scripts):
    output_files = []
    for script in scripts:
        script_path = os.path.join(working_directory, script["script"])
        try:
            print(f"Running script: {script_path}")
            subprocess.run(["python", script_path], check=True)
            output_files.append(os.path.join(working_directory, script["output"]))
        except subprocess.CalledProcessError as e:
            print(f"❌ Error running {script['script']}: {e}")
            continue
    return output_files

def copy_cell_styles(source_cell, target_cell):
    """
    Copies all styles from a source cell to a target cell.
    """
    if source_cell and target_cell:
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                vertAlign=source_cell.font.vertAlign,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color
            )
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
        if source_cell.fill:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text
            )
        target_cell.number_format = source_cell.number_format

def handle_merged_cells(ws_src, ws_dest):
    """
    Handles merged cells in the source worksheet and applies them to the destination worksheet.
    """
    for merged_range in ws_src.merged_cells.ranges:
        ws_dest.merge_cells(str(merged_range))
        # Apply styles to the top-left cell of the merged range
        top_left_cell = ws_src.cell(merged_range.min_row, merged_range.min_col)
        dest_top_left_cell = ws_dest.cell(merged_range.min_row, merged_range.min_col)
        copy_cell_styles(top_left_cell, dest_top_left_cell)

def combine_excel_files(file_paths, output_file):
    """
    Combines multiple Excel files into one while preserving all formats.
    Handles AS_EST_ORTH.xlsx separately due to format differences.
    """
    if not file_paths:
        print("❌ No files to combine.")
        return

    # Load the first workbook as the base
    wb_combined = load_workbook(file_paths[0])
    first_sheet = wb_combined.active
    ref_c_width = first_sheet.column_dimensions['C'].width if 'C' in first_sheet.column_dimensions else 5

    for file_path in file_paths[1:]:
        wb = load_workbook(file_path)
        file_name = os.path.basename(file_path).upper()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            new_sheet_name = sheet_name
            if new_sheet_name in wb_combined.sheetnames:
                new_sheet_name += "_copy"
            new_sheet = wb_combined.create_sheet(title=new_sheet_name)

            # Copy cell values and styles
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if cell is not None:
                        new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                        copy_cell_styles(cell, new_cell)

            # Handle merged cells
            handle_merged_cells(ws, new_sheet)

            # Copy column widths safely
            for col in ws.columns:
                try:
                    col_letter = get_column_letter(col[0].column)
                    if col_letter in ws.column_dimensions:
                        new_sheet.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width
                except Exception:
                    continue

            # Copy common properties
            new_sheet.page_setup = ws.page_setup
            new_sheet.sheet_properties = ws.sheet_properties
            new_sheet.page_margins = ws.page_margins
            if ws.sheet_view:
                new_sheet.freeze_panes = ws.freeze_panes

            # 🧠 Apply conditional formatting logic
            try:
                if "ORTH" not in file_name:
                    # Normal logic for CTC sheets
                    new_sheet.column_dimensions['A'].hidden = True

                    # Shrink column widths from column E onwards
                    for col in new_sheet.iter_cols(min_col=5):
                        col_letter = col[0].column_letter
                        new_sheet.column_dimensions[col_letter].width = 9.5
                        
                    protected_sheets = ["AS EST TOP 20 CTC", "DO EST TOP 20 CTC"]
                    
                    if new_sheet.title not in protected_sheets:
                    # Shrink column C specifically
                        for col in new_sheet.iter_cols(min_col=3, max_col=3):
                            col_letter = col[0].column_letter
                            new_sheet.column_dimensions[col_letter].width = 5
                            new_sheet.column_dimensions['C'].width = ref_c_width
                    else:
                        print(f"✔ Column C kept unchanged for sheet: {new_sheet.title}")

                    # Zoom out to 80%
                    new_sheet.sheet_view.zoomScale = 80
                else:
                    # Lighter formatting for ORTH sheet
                    new_sheet.sheet_view.zoomScale = 80
                    new_sheet.column_dimensions['A'].hidden = True
                    print(f"ℹ️ Skipped column resizing for {file_name}")
            except Exception as e:
                print(f"⚠️ Formatting error in sheet '{new_sheet.title}': {e}")

    wb_combined.save(output_file)
    print(f"✅ Combined file saved as: {output_file}")

# Run scripts and collect output files
output_files = run_scripts_and_collect_outputs(scripts)

# Combine the resulting Excel files
combine_excel_files(output_files, output_file)

'''
import smtplib
from email.message import EmailMessage
from email.utils import formataddr

# Email details
sender_email = "website@parcon.in"
sender_name = "Oushnik Sarkar"
receiver_name = "Kshitij Nilank"
#receiver_email = "mis@parcon-india.com"
#cc_emails = ["sebi@goodricke.com","pdey@parcon-india.com","mis@parcon-india.com"]
#bcc_emails = ["oushnik@gmail.com", "oushnik23@gmail.com"]

subject = "GGL EST AS/DO CTC & TOP 20 AS/DO EST CTC & ORTH GRADEWISE QTY & AVG. PRICE FOR & UPTO SALE - 46 SEASON 2025/26"
body = """<p>Dear Sir,</p>

<p>Please find the attached Excel file.</P>

<p><strong>Regards,</strong><br>
Oushnik Sarkar<br>
<strong>Parcon (India) Pvt. Ltd.</strong></p>"""

# Create email
msg = EmailMessage()
msg['Subject'] = subject
msg['From'] = f"{sender_name} <{sender_email}>"
msg['To'] = formataddr(("Kshitij Nilank", "mis@parcon-india.com"))
#msg['Cc']=cc_emails
#msg['Bcc'] = ", ".join(bcc_emails)
msg.add_alternative(body, subtype='html')

# Attach the Excel file
with open('GGL FOR SALE- 46.xlsx', 'rb') as f:
    file_data = f.read()
    file_name = 'GGL FOR SALE- 46.xlsx'

msg.add_attachment(file_data, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=file_name)

# SMTP server setup (Example for Gmail)
smtp_server = "smtp.gmail.com"
smtp_port = 587

# Send email
with smtplib.SMTP(smtp_server, smtp_port) as server:
    server.starttls()
    server.login(sender_email, "llys iazq ddtd ruyj")  # Replace with your real password
    server.send_message(msg)

print("✅ Email sent successfully!")'''