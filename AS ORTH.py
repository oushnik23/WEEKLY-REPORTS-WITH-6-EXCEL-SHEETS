#!pip install db-dtypes

#import sys
#print(sys.executable)

import pandas as pd
import datetime
import numpy as np
import mysql.connector as msql
from mysql.connector import Error
from google.cloud import bigquery
import warnings
warnings.filterwarnings("ignore")

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

import os 
os.chdir(r"D:\Oushnik Sarkar\Python\Weekly\GGL")
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = 'D:/Oushnik Sarkar/data-warehousing-prod.json'

#Imports google cloud client library and initiates BQ service
from google.cloud import bigquery
#from google.cloud import datastore
bigquery_client = bigquery.Client()

QUERY = """
select FinYear,
CASE 
    WHEN Area = "AS" AND Centre IN ("KOL", "GUW") THEN "AS"
    WHEN Area IN ("DO", "TR") AND Centre IN ("KOL", "SIL") THEN "DO/TR"
    WHEN Area IN ("CA", "TP") AND Centre IN ("KOL", "GUW") THEN "CA/TP" ELSE Area END AS AreaAlies,
case	
	WHEN SubTeaType = "PRIMARY" THEN "P"
	WHEN SubTeaType = "SECONDARY" THEN "S" ELSE SubTeaType END AS PS,
SaleNo,
if(SaleNo >=1 and SaleNo <=13, 52+SaleNo,SaleNo) as SalesAlies,
BrokerCode,Area,Category,GardenMDM,GradeMDM,
coalesce(Centre,"") as Centre,SellerGroup,
coalesce(EstBlf,"") as EstBlf, 
ROUND( SAFE_DIVIDE(SUM(Value) , SUM(TotalWeight)) ,2) AS AvgPrice,
#coalesce(sum(IF(LotStatus = 'Sold',TotalWeight,InvoiceWeight)),0) as OfferQty,
sum(TotalWeight) as Sold_Qty,
coalesce(sum(Value),0) as Total_Value

FROM  `data-warehousing-prod.EasyReports.SaleTransactionView`

where Area IN ("AS","DO","TR") and EstBlf IN ("EST") and Centre IN ("KOL","GUW","SIL") and Season IN (2025) and 
if(SaleNo >=1 and SaleNo <=13, 52+SaleNo,SaleNo) between 14 and (Select max(if(SaleNo>=1 AND SaleNo<=13,52+SaleNo,SaleNo)) from data-warehousing-prod.EasyReports.SaleTransactionView 
	where Season = 2025 and FinYear = "2025-26")

group by FinYear,SaleNo,SalesAlies,BrokerCode,Area,Category,GardenMDM,GradeMDM,PS,AreaAlies,Centre,SellerGroup,EstBlf
Having FinYear IS NOT NULL
"""
Query_Results = bigquery_client.query(QUERY)
df = Query_Results.to_dataframe()

#----------------------------------------------------------------------------------------

QUERY2 = """
SELECT FinYear, coalesce(SellerGroup,"") as SellerGroup,
Case 
when Area = "AS" and Centre IN("KOL","GUW") then "AS"
when Area IN ("DO","TR") and Centre IN("KOL","SIL") then "DO/TR" 
else "" end as AreaAlies,  
	Category,GardenMDM,
    Case when BrokerCode = "PC" then "PC" else "Other" end as BrokerCode,
    coalesce( ROUND(sum(Value),0),0) as Total_Value,
 coalesce(ROUND(sum(TotalWeight),0),0) as Sold_Qty
FROM `data-warehousing-prod.EasyReports.SaleTransactionView`
WHERE
  BrokerCode = "PC" AND FinYear = '2025-26' AND Season = 2025 and EstBlf = "EST" and Category IN("ORTHODOX") and 
  Centre IN("KOL","SIL","GUW") and if(SaleNo>=1 AND SaleNo<=13,52+SaleNo,SaleNo) between 14 and (Select max(if(SaleNo>=1 AND SaleNo<=13,52+SaleNo,SaleNo)) from data-warehousing-prod.EasyReports.SaleTransactionView 
	where Season = 2025 and FinYear = "2025-26") 
  And SellerGroup IN ('GOODRICKE')
	 
GROUP BY FinYear,SellerGroup,Category,GardenMDM , BrokerCode,AreaAlies
Having SellerGroup<>"" and sum(TotalWeight)>0.0000000001 and AreaAlies<>""
"""

Query_Results2 = bigquery_client.query(QUERY2)
df_pc = Query_Results2.to_dataframe()

df_pc['Avg']=df_pc['Total_Value'] / df_pc['Sold_Qty']
df_pc['BrokerCode'].unique()

#----------------------------------------------------------------------------------------
Sale = df['SalesAlies'].max() - 52 if df['SalesAlies'].max() > 52 else df['SalesAlies'].max()

df=df[(df['Category'].isin(['ORTHODOX'])) & (df['SalesAlies']==Sale) & (df['AreaAlies'].isin(['AS'])) & 
      (df['SellerGroup']=='GOODRICKE')]

summary=df.groupby(['GradeMDM','GardenMDM'],as_index=False).agg({'Sold_Qty':'sum','Total_Value':'sum'})

summary['Avg']=summary['Total_Value'] / summary['Sold_Qty']
summary.drop(['Total_Value'],inplace=True,axis=1)

#-----------------------------------------------------------------------------------------#
'''
summary1=df.groupby(['PS','GradeMDM','GardenMDM'],as_index=False).agg({'Sold_Qty':'sum','Total_Value':'sum'})
summary1['Avg']=summary1['Total_Value'] / summary1['Sold_Qty']
summary1.drop(['Total_Value'],inplace=True,axis=1)
summary1['Centre']='Comb'

column_to_move = summary1.pop("Centre")
summary1.insert(0, "Centre", column_to_move)'''

#----------------------------------------Combined-------------------------------------------------#
'''
comb=pd.concat([summary,summary1])
df2=comb.set_index(['Centre','PS','GradeMDM','GardenMDM']).unstack(['GardenMDM'])'''

df2=summary.set_index(['GradeMDM','GardenMDM']).unstack(['GardenMDM'])
df2=df2.swaplevel(axis=1).sort_index(axis=1)

#-----------------------------------------------------------------------------------------#

desired_order=['Sold_Qty','Avg']
new_columns = []
for i in df2.columns.levels[0]: 
    for j in desired_order:  
        new_columns.append((i,j))
df2 = df2[new_columns]

#--------------------------------------SUBTOTAL2---------------------------------------------------#

# Step 1: Initialize a dictionary to collect grand total values
grand_total_data = {}

# Step 2: Loop through each garden
for garden in df2.columns.levels[0]:
    qty = df2[(garden, 'Sold_Qty')]
    avg = df2[(garden, 'Avg')]

    total_qty = qty.sum()
    weighted_avg = (qty * avg).sum() / total_qty if total_qty != 0 else 0

    grand_total_data[(garden, 'Sold_Qty')] = total_qty
    grand_total_data[(garden, 'Avg')] = round(weighted_avg, 2)

# Step 3: Create a DataFrame for the Grand Total row
grand_total_df = pd.DataFrame(grand_total_data, index=['Grand Total'])
grand_total_df.index.name = 'GradeMDM'

# Step 4: Append to df2
df4 = pd.concat([df2, grand_total_df])

#--------------------------------------Desired Orders---------------------------------------------------#

# Define desired orders

grade_sequence = ["BPS","BOP","OF","BOP1","FTGFOP1(SPL)CL","FTGFOP1(CL)","FTGFOP1(CH)SPL","FTGFOP1(CL)SPL",
"SFTGFOP1(SPL)CL","SFTGFOP1","FTGFOP(CL)","STGFOP1(S)","STGFOP1","STGFOP","FTGFOP1","GTGFOP1","FTGFOP","TGFOP1(SPL)"
,"TGFOP1(CL)","TGFOP1","TGFOP(SPL)","TGFOP(CL)","TGFOP","TGFOPS","TGFOP(S)","GFOP","GFOP(SPL)","GFOP(CL)",
"FOP(CL)","FOP","BPS(CL)","OBPS","TBOP(CL)","TBOP","OBPS1","TFBOP(CL)","TFBOP","FBOP(CL)","FBOP","FBOPSUPREME",
"TGFBOP1(CL)","TGFBOP(SPL)","TGFBOP(CL)","TGFBOP","GFBOP(SPL)","GFBOP(CL)","GFBOP","GFBOP(S)","GFBOPS","GBOP(TIPPY)"
,"FTGBOP","TGBOP1(CL)","TGBOP(SPL)","TGBOP(CL)","TGBOP","GBOP(SPL)","GBOP","BOP(S)","BPS(CLO)","GOF","FOF","TGOF"
,"BOPF","FOF(CL)","GFOP1","GFOP(S)","GFOPSUPREME","OP","FOP1","FOP(S)","BPS1","BOPSUPREME","FBOP(S)","GBOP(S)"
,"GBOPS","FBOP1","GFBOP1","GBOP1","GOF1","FOF1","TGOF1","TGBOP1","GTGFOP","1ST FINE","2ND FINE","3RD FINE","COARSE"
,"S.FINE","FINE","LACHHA","FORESTMISTSTGFOP1S","OOLONG","LFINE","SFINE","TOF","GOF(S)","SOMI","SFINE1(S)","LFINES",
"OFS","GOLDENTIPS","GT FINE TP","SUPERFINE","OPS","OPS1","SFINE CL","WHITE TEA","MOGRA I CLONAL","MOGRA II CLONAL",
"SUPER FINE CLONAL","FINE CLONAL","SUPER MOGRA I","SUPER FINE PREMIUM","FINE PREMIUM","MOGRA I PREMIUM",
"MOGRA II PREMIUM","MIDNIGHT BLUE","O.S.T","PREMIUM MAZDANA","MOGRA","FTGFOP1-MIDNIGHT BLUE","STGBOP1",
"STGFOP1(SPL)","P126 VELVETYWHITE FEATHER","WHOLE LEAF","OF(S)","BOPF(SPL)","TGFOP1(S)","FBOPF","GBOP(SM)",
"OP1","BOP(O)","GBOP(CL)","BOP(CL)","BPS(SPL)","FTGFOP1(RUBY)","FTGFOP1 EXCEL","SILVER PEKOE NEEDLES","STGFOP(SPL)"
,"GOLDENPEARL","FINE I","GREEN GARNET","SILVERNEEDLE","OP-A","MASDANA","SILVER TEA","FBOP(SPL)","SUPER LACCHA"
,"FINE I (S)","FD","FTGFOP1(S)","BPS(O)","TFBOP(S)","BPS(S)","TGOF(S)","OP CL","CLONAL DELIGHT","BPS(O)","MOGRA I"
,"MOGRA II","SUPER MOGRA","YOUNGHYSON","FANNINGS","STGFOP1S","MOONLIGHT","MOONDROP","CLONALDELIGHT","MOONSHINE","FTGFOP1(CH)","WHITE NEEDLE","SUPER TWIST","BOPF1","GTGOF","BOPD","FLOWERY CLONAL","WONDER","SPRING DELIGHT","ENIGMA","SILVER TIPS","GTFANNINGS","SFTGFOP (CL)","SUPERFINE1","SPECIALITYTEA","GREY NEEDLE","EXOTICOOLONG","SFTGFOP","RUBY","CLONALMUSK","FTGFOP1(MUSCATEL)","MYSTIC","FTGFOP(SPL)","GTGBOP","GUNPOWDER","MIRCHI MOGRA","KATH MOGRA","BT","HAND CRAFTED OOLONG","MOUNTAIN MIST","HST","SFTFOP1","GTBPS","PRFINE","GREEN FINE","GD","FBOPTIPPY","MONGRASMALL","TOF1","FTGFOP1TIPPYCLONAL","GTGFOP(SPL)","GTFD"
]

grade_order = {g: i for i, g in enumerate(grade_sequence)}

# Convert index to DataFrame
idx_df = df4.index.to_frame(index=False)

# Map sorting ranks

idx_df['grade_rank'] = idx_df['GradeMDM'].map(grade_order).fillna(998)

# Force 'Total' to the end
idx_df.loc[idx_df['GradeMDM'] == 'Grand Total', 'grade_rank'] = 999

# Sort and rebuild MultiIndex
sorted_index = idx_df.sort_values(['grade_rank']) \
                     .drop(columns=['grade_rank']) \
                     .apply(tuple, axis=1)
                     
sorted_index = idx_df.sort_values(['grade_rank'])['GradeMDM'].tolist()

df4 = df4.loc[[g for g in sorted_index if g in df4.index]]

#---------------------------------------GRAND TOTAL------------------------------------------#

# Step 1: Extract Sold_Qty and Avg from df2
sold_qty = df4.xs('Sold_Qty', axis=1, level=1)
avg = df4.xs('Avg', axis=1, level=1)

# Step 2: Calculate total Sold_Qty and weighted Avg
total_sold_qty = sold_qty.sum(axis=1)
weighted_avg = (sold_qty * avg).sum(axis=1) / total_sold_qty

# Step 3: Combine into a new DataFrame
grand_total_df = pd.DataFrame({
    ('Grand_Total', 'Sold_Qty'): total_sold_qty,
    ('Grand_Total', 'Avg'): weighted_avg
})

# Step 4: Reindex to match df4
grand_total_df = grand_total_df.reindex(df4.index)

# Step 5: Append to df4
df4 = pd.concat([df4, grand_total_df], axis=1)
'''
# Step 6: Fill missing 'Grand_Total' for subtotal and total rows
# Loop through each Centre and PS group
for centre in df4.index.get_level_values('Centre').unique():
    for ps in df4.index.get_level_values('PS').unique():
        # Filter rows that are not 'Total' or '*'
        mask = (df4.index.get_level_values('Centre') == centre) & \
               (df4.index.get_level_values('PS') == ps) & \
               (~df4.index.get_level_values('GradeMDM').isin(['Total', '*']))
        group = df4.loc[mask]

        # Calculate subtotal
        subtotal_qty = group[('Grand_Total', 'Sold_Qty')].sum()
        subtotal_avg = (group[('Grand_Total', 'Sold_Qty')] * group[('Grand_Total', 'Avg')]).sum() / subtotal_qty

        # Assign to 'Total' row
        total_idx = (centre, ps, 'Total')
        if total_idx in df4.index:
            df4.loc[total_idx, ('Grand_Total', 'Sold_Qty')] = subtotal_qty
            df4.loc[total_idx, ('Grand_Total', 'Avg')] = subtotal_avg

# Step 7: Fill '*' level totals
for centre in df4.index.get_level_values('Centre').unique():
    mask = (df4.index.get_level_values('Centre') == centre) & \
           (~df4.index.get_level_values('GradeMDM').isin(['Total', '*']))
    group = df4.loc[mask]

    total_qty = group[('Grand_Total', 'Sold_Qty')].sum()
    total_avg = (group[('Grand_Total', 'Sold_Qty')] * group[('Grand_Total', 'Avg')]).sum() / total_qty

    star_idx = (centre, '*', 'Grand Total')
    if star_idx in df4.index:
        df4.loc[star_idx, ('Grand_Total', 'Sold_Qty')] = total_qty
        df4.loc[star_idx, ('Grand_Total', 'Avg')] = total_avg
'''
#--------------------------------------Arranging PC Gardens----------------------------------------------------#

gardenlist=df_pc['GardenMDM'][(df_pc['BrokerCode']=='PC') & (df_pc['AreaAlies']=='AS') 
                              & (df_pc['Category']=='ORTHODOX')].unique()

# Step 1: Extract the 'Avg' values from ('Comb', '*', 'Total') row
comb_star_total_avg = df4.loc[('Grand Total')].xs('Avg', axis=0, level=1)

# Step 2: Convert gardenlist to list if it's a NumPy array
#gardenlist = gardenlist.tolist()


gardenlist_filtered = [g for g in gardenlist if g in comb_star_total_avg.index]

# Step 3: Sort priority gardens by their Avg in ('Comb', '*', 'Total') row
#priority_sorted = comb_star_total_avg[gardenlist].sort_values(ascending=False).index.tolist()

priority_sorted = comb_star_total_avg[gardenlist_filtered].sort_values(ascending=False).index.tolist()

# Step 4: Sort remaining gardens by their Avg in the same row
remaining_gardens = [g for g in comb_star_total_avg.index if g not in gardenlist and g != 'Grand_Total']
remaining_sorted = comb_star_total_avg[remaining_gardens].sort_values(ascending=False).index.tolist()

# Step 5: Combine final garden order
final_garden_order = priority_sorted + remaining_sorted

# Step 6: Build column list in garden-wise pairs
new_columns = []
for garden in final_garden_order:
    new_columns.append((garden, 'Sold_Qty'))
    new_columns.append((garden, 'Avg'))

# Step 7: Add Grand_Total at the end
new_columns.append(('Grand_Total', 'Sold_Qty'))
new_columns.append(('Grand_Total', 'Avg'))

# Step 8: Reorder df4 columns
df4 = df4.loc[:, pd.MultiIndex.from_tuples(new_columns)]


#------------------------------------------INSERT BLANK SPACE------------------------------------------#
'''
targets = [('KOL', '*', 'Grand Total'), ('GUW', '*', 'Grand Total')]

# Prepare blank row with same columns
blank_row = pd.DataFrame([[None]*df4.shape[1]], columns=df4.columns)
blank_row.index = pd.MultiIndex.from_tuples([('', '', '')], names=df4.index.names)

# Insert blank rows after each target
for target in reversed(targets):  # Reverse to avoid shifting positions
    if target in df4.index:
        pos = df4.index.get_loc(target) + 1
        df4_top = df4.iloc[:pos]
        df4_bottom = df4.iloc[pos:]
        df4 = pd.concat([df4_top, blank_row, df4_bottom])'''

#--------------------------------------------DESIGN-------------------------------------------#

df4=df4.reset_index()

excel_path = "AS ORTH.xlsx"

with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
    df4.to_excel(writer, sheet_name="ORTH", startrow=3)
    workbook  = writer.book
    worksheet = writer.sheets["ORTH"]
    
    worksheet.set_column('A:A', None, None, {'hidden': True})

    worksheet.freeze_panes(5, 2)  

    for i, col in enumerate(df4.columns):
        col_label = f"{col[0]} {col[1]}" if isinstance(col, tuple) else str(col)
        worksheet.set_column(i + 3, i + 3, max(12, len(col_label)))
        
wb = load_workbook(excel_path)
ws = wb["ORTH"]

thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),
                     bottom=Side(style='thin'))

for row in ws.iter_rows(min_row=4):
    for cell in row:
        cell.border = thin_border

ws["B1"] = "GGL ASSAM EST ORTHODOX"
ws["B2"] = f"FOR SALE - {Sale}"
ws["B3"] = "SEASON 2025/26"

bold_font = Font(bold=True,color="FF0000")
ws["B1"].font = bold_font
ws["B2"].font = bold_font
ws["B3"].font = bold_font

ws.delete_rows(6)
       
import re
for col in ws.iter_cols(min_col=3, max_row=ws.max_row):
    for cell in col:
        try:
            if isinstance(cell.value, str):
                
                cleaned = cell.value.replace(',', '').replace(' ', '')
                cleaned = re.sub(r'[^0-9.\-]', '', cell.value) 
                
                if re.match(r'^-?\d+(\.\d+)?$', cleaned): 
                    cell.value = float(cleaned)

                if cleaned.replace('.', '', 1).isdigit(): 
                    cell.value = float(cleaned) 

            
            if isinstance(cell.value, (int, float)):
                
                cell.number_format = '#,##0;-#,##0'  # Format without decimals for display purposes

                # Internally retain decimal value for future reference
                if isinstance(cell.value, float):
                    # Keep the original value (store as float internally)
                    pass  # No action needed, value remains as float

        except Exception as e:
            print(f"Skipping cell {cell.coordinate}: {e}")
          
targets = [('Grand Total')]

# Define the font style for Grand Total
highlight_font = Font(bold=True, color="375623")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    grade  = row[1].value  # Column D

    if (str(grade).strip().upper()) in \
       [g.upper() for g in targets]:
        for cell in row:
            cell.font = highlight_font

ws["B4"] = "Garden"
ws["B5"] = "Grade"

# Define center alignment
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Assuming data starts at row 2
    for cell in row[2:]:  # Skip first 2 columns (A, B)
        cell.alignment = Alignment(horizontal='center')

#Hide 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Assuming data starts at row 4
    for cell in row[2:]:  # Skip columns A–D
        if cell.value == 0:
            cell.value = ""

# Shrink column widths
for col in ws.iter_cols(min_col=2):  # Skip columns A, B, C, D
    col_letter = col[0].column_letter
    ws.column_dimensions[col_letter].width = 9.5
  
for col in ws.iter_cols(min_col=2,max_col=2):  
    col_letter = col[0].column_letter
    ws.column_dimensions[col_letter].width = 12
    
# Zoom out to 80%
ws.sheet_view.zoomScale = 80
                
wb.save(excel_path)
print(f"✅ Excel file {excel_path} created successfully!")
        
#print(df4.columns)
#print(df4.index)